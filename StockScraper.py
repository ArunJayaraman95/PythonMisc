from openpyxl import Workbook
from openpyxl.styles import *
from openpyxl.worksheet.table import Table, TableStyleInfo
import bs4
import requests
from bs4 import BeautifulSoup as bs

tickerList = [
    "IMGN"]  # , "AAPL", "XOM", "E", "FORD", "KSS", "SIX", "PLAY", "A", "ABM", "ACH", "ACC", "AEO", "AFGH", "AGI", "BTG"]

columnLength = len(tickerList)

# define workbook
workbook = Workbook()
filePath = r"C:\Users\ripar\Desktop\StockScreenTest.xlsx"
workbook.save(filePath)

sheet = workbook['Sheet']
sheet.title = 'Main Page'
sheet = workbook['Main Page']

maxShares = workbook.create_sheet('Max Shares')


totalList = []


def stockMagic(ticker):
    mainList = [ticker]
    # set urls for function
    urlHomeFirst = "https://finance.yahoo.com/quote/"
    urlHomeSecond = "?p="
    urlHomeThird = "&.tsrc=fin-srch"

    def getHomeUrl(x):
        return str(urlHomeFirst + x + urlHomeSecond + x + urlHomeThird)

    r = requests.get(getHomeUrl(ticker))
    r = requests.get(getHomeUrl(ticker))
    soup = bs(r.text, 'html.parser')

    priceData = soup.find('div', {'class': 'My(6px) Pos(r) smartphone_Mt(6px)'})
    price = priceData.find('span', {'class': 'Trsdu(0.3s) Trsdu(0.3s) Fw(b) Fz(36px) Mb(-4px) D(b)'}).text
    # priceChange = priceData.find('span', {'class': 'Trsdu(0.3s) Fw(500) Fz(14px) C($negativeColor)'}).text
    # if priceChange is None:
    #     priceChange = price.find('span', {'class': 'Trsdu(0.3s) Fw(500) Fz(14px) C($positiveColor)'}).text
    # print(price)

    # print(priceChange)

    statsUrl = "https://finance.yahoo.com/quote/" + ticker + "/key-statistics?"
    statsR = requests.get(statsUrl)
    statsSoup = bs(statsR.text, 'html.parser')
    table = statsSoup.findAll('td')

    testList = []
    for i in table:
        testList.append(i.text)

    while testList[0] != "Beta (5Y Monthly) ":
        del testList[0]

    # print(testList)

    def getValueStats(x):
        return testList[testList.index(x) + 1]

    def findValue(x):
        testList.index(x)
        # print(x + ": " + getValueStats(x))

    # findValue("Current Ratio (mrq)")
    tableHome = soup.findAll('td')
    homeList = []

    for i in tableHome:
        homeList.append(i.text)

    # print(homeList)

    def getValueHome(x):
        return homeList[homeList.index(x) + 1]

    mainList.append(price)
    # mainList.append(priceChange)
    mainList.append(getValueHome("Day's Range"))
    mainList.append(getValueHome("EPS (TTM)"))
    mainList.append(getValueHome("PE Ratio (TTM)"))
    mainList.append(getValueStats("Operating Margin (ttm)"))
    mainList.append(getValueStats("Quarterly Revenue Growth (yoy)"))
    mainList.append(getValueStats("Total Debt/Equity (mrq)"))
    totalList.append(mainList)

    print("Processed ", ticker)


# Append the label's onto the top

mainLabels = ["Name", "Price", "Day's Range", "EPS", "P/E Ratio", "Operating Margin", "Quart. Rev. Growth", \
              "Debt/Equity Ratio"]
sheet.append(mainLabels)


for tick in tickerList:
    stockMagic(tick)

# cycle through and append
for row in totalList:
    sheet.append(row)

# format column widths
columnList = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']
for col in columnList:
    sheet.column_dimensions[col].width = 20


def fc(x, y):  # quick def to assign a with to a column
    sheet.column_dimensions[x].width = y


fc('G', 18)

# Creating a table inside the sheet
table = Table(displayName="Table", ref="A1:H%d" % (len(tickerList) + 1))

# Defining a style for the table (default style name, row/column stripes)
# Choose your table style from the default styles of openpyxl
# Just type in openpyxl.worksheet.table.TABLESTYLES in the Python interpreter
style = TableStyleInfo(name="TableStyleMedium9", showRowStripes=True, showColumnStripes=False)

my_red = colors.Color(rgb='00FF0000')
my_fill = fills.PatternFill(patternType='solid', fgColor=my_red)
highFont = Font(color=colors.GREEN, bold=True)

redFill = PatternFill(start_color='00FFFFFF',
                      end_color='FFFF0000',
                      fill_type='darkDown')

# The range plus two is to adjust for the offset from the title and the index
# for cellX in range(2, columnLength + 2):
#
#     if float(sheet['B%s' % cellX].value) < 20:

# sheet['B%s' % cellX].font = highFont
# sheet['B%s' % cellX].fill = redFill
# Applying the style to the table
table.tableStyleInfo = style

# Adding the newly created table to the sheet
sheet.add_table(table)

workbook.save(filePath)
workbook.close()
